"""Secure, read-only XLSX inspection and normalization."""

from __future__ import annotations

from pathlib import Path, PurePosixPath
from typing import Mapping
import zipfile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .base import (
    DatasetImporter,
    FieldMapping,
    ImportInspection,
    ImportOptions,
    NormalizedDataset,
    PreviewPage,
    SheetInfo,
    StoredUpload,
    normalize_column_name,
    preview_from_records,
    write_normalized_dataset,
)


_REQUIRED_VALUE_TARGETS = frozenset(
    {"query", "text", "relevance", "shown_doc_text", "user_verdict"}
)
_MAPPING_ALIASES = {
    "domain": "domain",
    "query id": "query_id",
    "qid": "query_id",
    "query": "query",
    "question": "query",
    "query text": "query",
    "doc id": "doc_id",
    "shown doc id": "shown_doc_id",
    "id": "doc_id",
    "text": "text",
    "document": "text",
    "document text": "text",
    "relevance": "relevance",
    "label": "relevance",
    "split": "split",
    "rank": "rank",
    "score": "score",
    "case id": "case_id",
    "shown document": "shown_doc_text",
    "shown doc text": "shown_doc_text",
    "user verdict": "user_verdict",
    "verdict": "user_verdict",
}


def _mapping_key(value: str) -> str:
    return " ".join(value.replace("_", " ").strip().lower().split())


class XlsxImporter:
    formats = ("xlsx",)

    def __init__(
        self,
        output_dir: Path,
        *,
        max_uncompressed_bytes: int = 512 * 1024 * 1024,
        max_sheets: int = 32,
        max_rows: int = 250_000,
        max_columns: int = 256,
        max_zip_members: int = 10_000,
    ) -> None:
        self.output_dir = Path(output_dir).expanduser().resolve()
        for name, value in (
            ("max_uncompressed_bytes", max_uncompressed_bytes),
            ("max_sheets", max_sheets),
            ("max_rows", max_rows),
            ("max_columns", max_columns),
            ("max_zip_members", max_zip_members),
        ):
            if isinstance(value, bool) or not isinstance(value, int) or value <= 0:
                raise ValueError(f"{name} must be a positive integer")
        self.max_uncompressed_bytes = max_uncompressed_bytes
        self.max_sheets = max_sheets
        self.max_rows = max_rows
        self.max_columns = max_columns
        self.max_zip_members = max_zip_members

    def inspect(self, upload: StoredUpload) -> ImportInspection:
        workbook = self._open(upload)
        try:
            sheets = tuple(
                SheetInfo(
                    name=sheet.title,
                    visible=sheet.sheet_state == "visible",
                    rows=sheet.max_row,
                    columns=sheet.max_column,
                )
                for sheet in workbook.worksheets
            )
        finally:
            workbook.close()
        return ImportInspection("xlsx", (), 0, sheets=sheets)

    def preview(self, upload: StoredUpload, options: ImportOptions) -> PreviewPage:
        columns, records = self._read(upload, options)
        return preview_from_records(columns, records)

    def normalize(
        self,
        upload: StoredUpload,
        mapping: FieldMapping,
        options: ImportOptions | None = None,
    ) -> NormalizedDataset:
        _, records = self._read(upload, options or ImportOptions())
        for source, target in mapping.source_to_target.items():
            if target not in _REQUIRED_VALUE_TARGETS:
                continue
            for row_number, record in enumerate(records, start=1):
                value = record.get(source)
                if value is None or (isinstance(value, str) and not value.strip()):
                    raise ValueError(
                        "XLSX required column "
                        f"{source!r} has no cached value at data row {row_number}"
                    )
        return write_normalized_dataset(self.output_dir, upload, records, mapping)

    @staticmethod
    def suggest_mapping(columns: tuple[str, ...]) -> FieldMapping:
        mapping: dict[str, str] = {}
        assigned: set[str] = set()
        for source in columns:
            target = _MAPPING_ALIASES.get(_mapping_key(source))
            if target is not None and target not in assigned:
                mapping[source] = target
                assigned.add(target)
        return FieldMapping(mapping)

    def _open(self, upload: StoredUpload):
        self._preflight(upload)
        try:
            workbook = load_workbook(
                upload.source_path,
                read_only=True,
                data_only=True,
                keep_links=False,
            )
        except (InvalidFileException, OSError, zipfile.BadZipFile) as exc:
            raise ValueError("XLSX workbook cannot be opened safely") from exc
        if len(workbook.worksheets) > self.max_sheets:
            workbook.close()
            raise ValueError("XLSX workbook exceeds the worksheet limit")
        for sheet in workbook.worksheets:
            if sheet.max_row > self.max_rows:
                workbook.close()
                raise ValueError(f"XLSX worksheet {sheet.title!r} exceeds the row limit")
            if sheet.max_column > self.max_columns:
                workbook.close()
                raise ValueError(f"XLSX worksheet {sheet.title!r} exceeds the column limit")
        return workbook

    def _preflight(self, upload: StoredUpload) -> None:
        if upload.source_path.suffix.lower() != ".xlsx":
            raise ValueError("only .xlsx workbooks are supported")
        try:
            with zipfile.ZipFile(upload.source_path) as archive:
                members = archive.infolist()
        except (OSError, zipfile.BadZipFile) as exc:
            raise ValueError("XLSX must be a valid ZIP workbook") from exc
        if len(members) > self.max_zip_members:
            raise ValueError("XLSX workbook has too many ZIP members")
        uncompressed_bytes = 0
        for member in members:
            parts = PurePosixPath(member.filename).parts
            if (
                member.filename.startswith(("/", "\\"))
                or "\\" in member.filename
                or ".." in parts
                or "\x00" in member.filename
            ):
                raise ValueError("XLSX contains an unsafe ZIP member name")
            uncompressed_bytes += member.file_size
            if uncompressed_bytes > self.max_uncompressed_bytes:
                raise ValueError("XLSX exceeds the uncompressed size limit")

    def _read(
        self,
        upload: StoredUpload,
        options: ImportOptions,
    ) -> tuple[tuple[str, ...], list[Mapping[str, object]]]:
        workbook = self._open(upload)
        try:
            sheet = self._select_sheet(workbook, options)
            header_values = next(
                sheet.iter_rows(
                    min_row=options.header_row,
                    max_row=options.header_row,
                    values_only=True,
                ),
                (),
            )
            if not header_values:
                raise ValueError("XLSX selected header row is empty")
            headers = tuple(normalize_column_name(value) for value in header_values)
            if len(set(headers)) != len(headers):
                raise ValueError("XLSX header contains duplicate fields")
            records: list[Mapping[str, object]] = []
            for values in sheet.iter_rows(
                min_row=options.header_row + 1,
                values_only=True,
            ):
                record = {headers[index]: values[index] for index in range(len(headers))}
                if any(value is not None and str(value).strip() for value in record.values()):
                    records.append(record)
            if not records:
                raise ValueError("XLSX selected worksheet has no data rows")
            return headers, records
        finally:
            workbook.close()

    @staticmethod
    def _select_sheet(workbook, options: ImportOptions):
        if options.sheet_name is not None:
            if options.sheet_name not in workbook.sheetnames:
                raise ValueError(f"XLSX worksheet does not exist: {options.sheet_name}")
            return workbook[options.sheet_name]
        for sheet in workbook.worksheets:
            if sheet.sheet_state == "visible":
                return sheet
        raise ValueError("XLSX workbook has no visible worksheet")


assert isinstance(XlsxImporter(Path(".")), DatasetImporter)


__all__ = ["XlsxImporter"]
