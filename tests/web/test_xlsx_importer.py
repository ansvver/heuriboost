from __future__ import annotations

from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook

from heuriboost_rag.importers.base import FieldMapping, ImportOptions, StoredUpload
from heuriboost_rag.importers.xlsx_importer import XlsxImporter


class XlsxImporterTests(unittest.TestCase):
    def _workbook(self, path: Path) -> None:
        workbook = Workbook()
        visible = workbook.active
        visible.title = "Production cases"
        visible.append(["Ignored heading"])
        visible.append(["Question", "Shown document", "Verdict"])
        visible.append(["What is a wash sale?", "A wash sale has a 30-day rule.", "good"])
        hidden = workbook.create_sheet("Internal notes")
        hidden.sheet_state = "hidden"
        hidden.append(["secret"])
        workbook.save(path)
        workbook.close()

    def test_inspects_visible_and_hidden_sheets_and_uses_visible_sheet_by_default(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "cases.xlsx"
            self._workbook(path)
            importer = XlsxImporter(Path(tmp) / "datasets")
            upload = StoredUpload("xlsx", path, "cases.xlsx")

            inspection = importer.inspect(upload)
            self.assertEqual(
                [(sheet.name, sheet.visible) for sheet in inspection.sheets],
                [("Production cases", True), ("Internal notes", False)],
            )
            preview = importer.preview(upload, ImportOptions(header_row=2))
            self.assertEqual(preview.columns, ("Question", "Shown document", "Verdict"))
            self.assertEqual(preview.rows[0]["Question"], "What is a wash sale?")
            self.assertEqual(
                importer.suggest_mapping(preview.columns).source_to_target,
                {
                    "Question": "query",
                    "Shown document": "shown_doc_text",
                    "Verdict": "user_verdict",
                },
            )

    def test_normalization_rejects_missing_cached_formula_in_required_column(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "formula.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["query", "text", "relevance"])
            sheet.append(["What is a wash sale?", "=CONCAT(\"cached\", \" answer\")", "good"])
            workbook.save(path)
            workbook.close()
            importer = XlsxImporter(Path(tmp) / "datasets")

            with self.assertRaisesRegex(ValueError, "cached value"):
                importer.normalize(
                    StoredUpload("formula", path, "formula.xlsx"),
                    FieldMapping({"query": "query", "text": "text", "relevance": "relevance"}),
                )

    def test_rejects_macro_extensions_and_zip_size_limit_before_opening_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            macro_path = root / "unsafe.xlsm"
            macro_path.write_bytes(b"not an xlsx")
            with self.assertRaisesRegex(ValueError, "xlsx"):
                XlsxImporter(root / "datasets").inspect(
                    StoredUpload("macro", macro_path, "unsafe.xlsm")
                )

            workbook_path = root / "large.xlsx"
            self._workbook(workbook_path)
            with self.assertRaisesRegex(ValueError, "uncompressed"):
                XlsxImporter(root / "datasets", max_uncompressed_bytes=1).inspect(
                    StoredUpload("large", workbook_path, "large.xlsx")
                )


if __name__ == "__main__":
    unittest.main()
